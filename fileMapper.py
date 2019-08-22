"""
Project Summary:

The aim is to convert all file names from their ID
to a title, which is given in an ExCeL spreadsheet.

The steps I need to take:
    --Given a file name in our Directory, find the matching ID1 in the spreadsheet ✓
    --Given ID1, get the title ✓
    --Open the file labelled ID1 and the file labelled ID2 underneath it ✓
    --Rename all files in Folder_ID2 to Title + ID2 ✓
    --Copy the renamed file somewhere else? A new directory ✓

Some checks I need to make:
    --What if the file name doesn't match an ID? ✓
    --Check all file IDs are unique ✓
    --That the file is a PDF (Folder Kind = Adobe Acrobat) ✓
    --Flag duplicates ✓
In each case there is some error, flag the error (in a .txt file?)
    --Create a textfile that documents what the python script is doing
    --Error handling ✓

To Do List:
    --Write the ID2s into the spreadsheet with ID1s and filenames ✓
        >In order to do this I need to move from xlrd to openpyxl to enable reading AND writing ✓
    --Allow user input for ID_col, Title_col and kind_col ✓
    --Progress bar ✓
    --Better handling of duplicate files ✓
"""


def removeDuplicateFiles(directory, duplicateBin):
    #Takes care of removing duplicate files, places them in a "bin" in the flat directory
    for files in os.listdir(directory):
        duplicateId = files.split(".")[0].strip()
        #Duplicates are of the form "ID (int)" which is not an integer.
        #Non-duplicates are of the form "ID" which is an integer
        if not duplicateId.isdigit():
            os.rename(directory + "//" + files, duplicateBin + "//" + files)

def checkIfUnique(sheet, column, start_row): #Just a check to make sure that the IDs uniquely identify the file's title
    for col in sheet.iter_rows(min_row = start_row, min_col = column, max_col = column):
        for cell in col:
            checkAgainst = cell.value
            for col2 in sheet.iter_rows(min_row = cell.row + 1, min_col = column, max_col = column):
                for cell2 in col2:
                    if cell2.value == checkAgainst: return False
    return True

def searchCol(sheet, column, start_row, value): #Searches for value in a given column in a given sheet below a certain row.
    for col in sheet.iter_rows(min_row = start_row, min_col = column, max_col = column):
        for cell in col:
            if str(cell.value) == value: return cell.row
    return -1

def get_immediate_subdirectories(a_dir):#Gets the first-order subdirectories
    #Credit to: https://stackoverflow.com/a/800201/6363186
    return [name for name in os.listdir(a_dir)
            if os.path.isdir(os.path.join(a_dir, name))]

def renameFiles(path, spreadsheetLocation, directory, holdingFolderName, sheetName, folderType, start_row, ID_col, title_col, kind_col, ID2_col):
    #Logs keep info on what is happening whilst folders are being renamed.
    errorLog = "\nERRORS:"
    concernLog = "\n\nCONCERNS:"
    successLog = "\n\nSUCCESSES:"

    total_files = 0
    renamed_files = 0

    for r, d, f in os.walk(directory):
        for files in f:
            total_files += 1

    try:
        #Opening the Excel document at the requested sheet
        workbook = load_workbook(spreadsheetLocation)#xlrd.open_workbook(spreadsheetLocation, on_demand = True)
        sheet = workbook[sheetName] #workbook.sheet_by_name(sheetName)

        #Checking to make sure the ID uniquely identifies the file, if IDs aren't unique, raise an error
        if not checkIfUnique(sheet, ID_col, start_row): raise Exception("Data type does not unique identify file") #check_col_is_unique(sheet, ID_col): raise Exception("Data type does not unique identify file")

        #Holding folder holds all the files once renamed in a flat structure
        holding_folder = os.mkdir(path + "\\" + holdingFolderName)
        duplicate_bin = os.mkdir(path + "\\" + holdingFolderName + "\\" + "Duplicate Bin")

        #The main renaming code is here:
        for folder in get_immediate_subdirectories(directory):
            row = searchCol(sheet, ID_col, start_row, folder)
            if row > -1: #i.e. if our search has given valid results
                ID = sheet.cell(row = row, column = ID_col).value
                kind = sheet.cell(row = row, column = kind_col).value
                title = sheet.cell(row = row, column = title_col).value
                #sheet.cell_value(row, ID_col), sheet.cell_value(row, kind_col), sheet.cell_value(row, title_col)
                #If the ID corresponds to the wanted kind, or no kind was inputted, we proceed to rename the files in the directory
                if(kind == folderType or not folderType):
                    subdirectory = directory + "\\" + folder
                    #Due to how the directories are structured, the files are located in the next level subdirectory
                    for folder in os.listdir(subdirectory):
                        subsubdir = subdirectory + "\\" + folder
                        if len(os.listdir(subsubdir)) > 1: #This if-statement flags up any duplicate files
                            concernLog += "\nRemoving duplicate files in %s" % subsubdir
                            renamed_files += len(os.listdir(subsubdir)) - 1
                            removeDuplicateFiles(subsubdir, path + "\\" + holdingFolderName + "\\" + "Duplicate Bin")
                        elif len(os.listdir(subsubdir)) == 0:
                            concernLog += "\nDirectory %s is empty" % subsubdir
                        for filename in os.listdir(subsubdir): #Here the renaming/duplication is handled
                            firstNewName = title + "_" + filename
                            oldName = subsubdir + "\\" + filename
                            newName = subsubdir + "\\" + firstNewName
                            os.rename(oldName, newName) #Handle renaming files
                            copyfile(newName, path + "\\" + holdingFolderName + "\\" + firstNewName) #handling copying files
                            successLog += "\n%s successfully renamed to %s" %(filename, firstNewName)
                            sheet.cell(row = row, column = ID2_col).value = filename
                            renamed_files += 1
                else: #Handle cases when the file isn't of the wanted folder type
                    concernLog += "\nID %s is of kind %s" % (ID, kind)
                    renamed_files += 1
            else:
                concernLog += "\nThe file %s has no corresponding entry in the spreadsheet in column %d" %(folder, ID_col)
                renamed_files += 1
            #Creation of and updating the progress bar
            progress = 100 * (renamed_files/total_files)
            bar = ("█" * int(progress)) + ("-" * (100 - int(progress)))
            sys.stdout.write("\rFile Renaming Progress: %s %6.2f%%" %(bar, progress))
            sys.stdout.flush()
        workbook.save(spreadsheetLocation)
    except Exception as e:
        errorLog += "\n%s\n" % e
    finally:
        errorLog += "\nNo more errors found during renaming"
        concernLog += "\nNo more concerns to be raised during renaming"
        successLog += "\nFinished"
        print(errorLog)
        print(concernLog)
        print(successLog)
        completeLog = open(path + "\\" + holdingFolderName + "\\" + "FileMappingLog.txt", "w+") #Makes a textfile of the log and saves it to flat directory
        completeLog.write(errorLog + concernLog + successLog)
        print("The information in the terminal is saved in FileMappingLog.txt in the Flat Directory.")
        completeLog.close()

def main():#Executes the main code
    print("Select the directory holding all the files to be renamed")
    directory = askdirectory()
    print("The directory selected to hold the files is %s" % directory)
    print("Select the directory that will hold the flat directory.")
    path = askdirectory()
    print("The directory selected to hold the flat directory is %s" % path)
    print("Select the spreadsheet with the IDs and file titles.")
    print("WARNING: DO NOT LEAVE THE SPREADSHEET OPEN WHILE THE PROGRAM RUNS.")
    spreadsheetLocation = askopenfilename()
    print("The spreadsheet selected is %s" % spreadsheetLocation)
    sheetName = input("Enter the name of the sheet with the ID and title data: ")
    folderType = input("Enter the kind of files you want renamed (leave blank if you want them all renamed): ")
    start_row = int(input("Enter the number giving the first row of data (first row of the spreadsheet not including headings): "))
    holding_folder_name = "Flat Directory" + str(random.randint(0,10000))
    ID_col = int(input("Enter the column (as a number) holding the IDs (e.g. for A=1, B=2, ... AA = 27, etc.): ")) #The column for IDs
    title_col = int(input("Enter the column (as a number) holding the titles (e.g. for A=1, B=2, ... AA = 27, etc.): ")) #The column for titles
    kind_col = int(input("Enter the column (as a number) holding the kind (e.g. for A=1, B=2, ... AA = 27, etc.): ")) #The column for the kind
    ID2_col = int(input("Enter the column (as a number) that you want to store the ID2s (pick an empty column to avoid overwriting any data): "))
    renameFiles(path, spreadsheetLocation, directory, holding_folder_name, sheetName, folderType, start_row, ID_col, title_col, kind_col, ID2_col)

try:
    from openpyxl import load_workbook
    from os import listdir
    import os
    import sys
    from shutil import copyfile
    from tkinter.filedialog import askdirectory, askopenfilename
    import random
    main()
except ImportError as e:
    print(e)